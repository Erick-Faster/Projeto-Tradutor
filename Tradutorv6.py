# -*- coding: utf-8 -*-
"""
Created on Wed May  1 19:27:07 2019

@author: Faster-PC
"""

'''
Pegando palavras do excel
'''

import openpyxl, os, pyperclip
import pandas as pd


def ColetaXls():
    x = [] #Vetor que sera inserido as palavras lidas
    os.chdir('C:\\Users\\Faster-PC\\MyPythonFiles') #Pasta de Trabalho

    workbook = openpyxl.load_workbook('DadosW.xlsx') #Abrir essa planilha
    type(workbook) # converte em tipo

    workbook.get_sheet_names() #Mostra os nomes das abas
    sheet = workbook.get_sheet_by_name('Plan1') #Abre a aba
    type(sheet) #converte aba pra tipo

    for i in range (1,50): #Colocar limite + 1
        x.append(sheet.cell(row=i,column=1).value) #Add i linhas da coluna 1 no vetor
        
    return x

def ColetaCsv():
    base = pd.read_csv('DadosCC.csv')
    x = base.iloc[:,0]
    x = list(x)
    return x

def conversao():
    converter = int(input("Deseja ler o arquivo em: [1]Xls ou [2]Csv? : "))
    if converter == 1:
        palavras = ColetaXls()
    elif converter == 2:
        palavras = ColetaCsv()
    else:
        palavras = [pyperclip.paste(),'exemplo','queijo']
    return palavras

def ConvIdioma():
    x = int(input("Para qual idioma?\n[1]Alemao ou [2]Ingles? : "))
    if x == 1:
        idioma = 'portugiesisch-deutsch'
    else:
        idioma = 'portugiesisch-englisch'
    return idioma

palavras = conversao() #Decide se ira converter xsl ou csv
idioma = ConvIdioma() #Decide o idioma
        
'''
Coletando informacoes e convertendo vetor
'''

from selenium import webdriver

browser = webdriver.PhantomJS() #Roda navegador fantasma (req instalacao)
frasesfinais = []

for i in range (49): #Colocar limite exato   
    browser.get("https://context.reverso.net/%C3%BCbersetzung/"+idioma+"/"+palavras[i]) #site no qual informacao eh extraida
    frase = browser.find_elements_by_class_name('text') #Pega TODOS os elementos de classe 'text' no site
    traducao = browser.find_elements_by_class_name('translation')
    try:
        tipo = browser.find_element_by_class_name('n')
        tipo = tipo.text
        if tipo == 'Substantiv':
            try:
                tipo_aux = browser.find_element_by_class_name('v')
                tipo = 'Substantiv (oder Verb)'
                print('\nTipo[%s] = Sub + Verb'%palavras[i])
            except:
                tipo = 'Substantiv'
                print('\nTipo[%s] = Sub X'%palavras[i])
                
    except:
        try:
            tipo = browser.find_element_by_class_name('v')
            tipo = tipo.text
            print('\nTipo[%s] = Verb'%palavras[i])
        except:
            tipo = 'Outros'
            print('\nTipo[%s] = Outros'%palavras[i])
    
    
    frasesprontas = [] #Cria vetor no qual ira as palavras formatadas
    traducoesprontas = []
    
    #Converte dados das frases
    for j in range (len(frase)):
        frasesprontas.append(frase[j].text) #formata arquivo html em string

    #Converte dados das traducoes
    for j in range (len(traducao)):
        traducoesprontas.append(traducao[j].text)

    #Limpa vazios das frases    
    for j in range(len(frasesprontas)):
        try:
            frasesprontas.remove("") #Remove todos os vazios da string
        except:
            print('%s Processado =)'%palavras[i]) #Surge quando todos os vazios sumirem
            break

    #Limpa vazios das traducoes
    for j in range(len(traducoesprontas)):
        try:
            traducoesprontas.remove("") #Remove todos os vazios da string
        except:
            print('%s Limpo =)'%palavras[i]) #Surge quando todos os vazios sumirem
            break
        
    #Alimenta vetor final que ira para o excel
    if frasesprontas[0] == 'Meinst Du:': #Corrige se palavra apresentar typo 
        if len(traducoesprontas) > 1: #Se tiver mais de uma sugestao de traducao
            frasesfinais.append([palavras[i],traducoesprontas[0]+', '+traducoesprontas[1],' ', frasesprontas[2] + ' | ' +frasesprontas[1],tipo]) #cria vetor final
        else: #Se houver somente uma sugestao de traducao
            frasesfinais.append([palavras[i],traducoesprontas[0],' ', frasesprontas[2] + ' | ' +frasesprontas[1],tipo.text]) #cria vetor final
    else:
        if len(traducoesprontas) > 1:
            frasesfinais.append([palavras[i],traducoesprontas[0]+', '+traducoesprontas[1],' ', frasesprontas[1] + ' | ' +frasesprontas[0],tipo]) #cria vetor final
        else:
            frasesfinais.append([palavras[i],traducoesprontas[0],' ', frasesprontas[1] + ' | ' +frasesprontas[0],tipo]) #cria vetor final
    
browser.close()
'''
Salvando o arquivo no excel
'''

workbook = openpyxl.Workbook() #Cria uma nova planilha
workbook.get_sheet_names()
sheet = workbook.get_sheet_by_name('Sheet')

#Adiciona vetor ao excel, linha por linga
for i in range (len(frasesfinais)):
    workbook.active.append(frasesfinais[i])

#Salva o arquivo excel
os.chdir('C:\\Users\\Faster-PC\\MyPythonFiles')
workbook.save('DadosW5.xlsx')

